import csv
import json
import logging
from typing import Iterable, Tuple, Dict, Any
from django.http import HttpResponse
from django.utils import timezone
from .models import Programa, Disertante, PostulacionDisertante, Edicion
from .services import sync_postulacion_a_disertante

logger = logging.getLogger('django.services_programa')


def _mapear_categoria(ejes_raw: str) -> str:
    """Mapea los ejes temáticos seleccionados en la postulación o planilla a las categorías del Programa."""
    if not ejes_raw or str(ejes_raw).strip() in ('—', '-', '', 'None'):
        return "LOGISTICA"
    
    ejes_lower = str(ejes_raw).lower()
    
    if "puerto" in ejes_lower or "comex" in ejes_lower or "comercio" in ejes_lower or "transporte" in ejes_lower or "movilidad" in ejes_lower:
        return "PUERTOS/COMERCIO EXTERIOR"
    elif "tecnolog" in ejes_lower or "ia" in ejes_lower or "digital" in ejes_lower or "data" in ejes_lower or "wms" in ejes_lower:
        return "TECNOLOGIA"
    elif "e-com" in ejes_lower or "comercio electronico" in ejes_lower:
        return "E-COMMERS"
    elif "supply" in ejes_lower or "cadena" in ejes_lower or "operaciones" in ejes_lower or "lean" in ejes_lower:
        return "SUPPLY CHAIN"
    elif "humano" in ejes_lower or "talento" in ejes_lower or "recursos" in ejes_lower or "management" in ejes_lower or "persona" in ejes_lower or "estrategia" in ejes_lower:
        return "CAPITAL HUMANO"
    elif "radio" in ejes_lower:
        return "RADIO"
    elif "sustentab" in ejes_lower or "verde" in ejes_lower or "ecolog" in ejes_lower or "innovacion" in ejes_lower or "offshore" in ejes_lower or "caso" in ejes_lower:
        return "SUSTENTABILIDAD"
    elif "apertura" in ejes_lower or "cierre" in ejes_lower or "institucional" in ejes_lower or "general" in ejes_lower:
        return "INSTITUCIONAL"
    elif "coffee" in ejes_lower or "networking" in ejes_lower:
        return "NETWORKING"
    elif "taller" in ejes_lower or "demo" in ejes_lower or "hackaton" in ejes_lower:
        return "WORKSHOP"
    else:
        return "LOGISTICA"


def _mapear_aula(aula_raw: str) -> str:
    """Normaliza nombres de aula como 'HUB 1 - Aula 1' a 'Aula 1' o devuelve el nombre limpio si es un aula dinámica."""
    if not aula_raw or str(aula_raw).strip() in ('—', '-', '', 'None'):
        return "Aula Magna"
    
    a_str = str(aula_raw).strip()
    a_lower = a_str.lower()
    
    for num in range(1, 30):
        if f"aula {num}" in a_lower or f"aula 0{num}" in a_lower:
            return f"Aula {num}"
            
    if "magna" in a_lower:
        return "Aula Magna"
        
    return a_str


def _mapear_estado(estado_raw: str) -> str:
    """Mapea estados como 'Bloqueado', 'Confirmado', 'Programado' a 'PUBLICADO' o 'BORRADOR'."""
    if not estado_raw or str(estado_raw).strip() in ('—', '-', '', 'None'):
        return "PUBLICADO"
    est_lower = str(estado_raw).strip().lower()
    if est_lower in ("bloqueado", "borrador", "draft"):
        return "BORRADOR"
    return "PUBLICADO"


def analizar_y_procesar_excel_programa(
    file_obj, 
    edicion=None, 
    commit=False, 
    filas_aprobadas=None,
    modo='SMART_SYNC'
) -> Dict[str, Any]:
    """
    Analiza un archivo Excel (.xlsx o .csv) y ejecuta el Pre-flight Analysis / Smart Delta Sync.
    Si commit=False, realiza un análisis en memoria (dry-run) clasificando cada fila en:
      - SIN_CAMBIOS (⚪)
      - MODIFICADO (🟡) con diff
      - NUEVO (🟢)
      - ERROR (🔴) con mensaje
    Si commit=True, guarda únicamente las filas especificadas en `filas_aprobadas` (o todas las válidas si None).
    """
    import openpyxl
    from datetime import datetime, time
    from django.db import transaction

    if edicion is None:
        edicion = Edicion.objects.filter(activa=True).first()

    # Cargar workbook
    rows_data = []
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows_data.append(list(row))
    except Exception as e:
        logger.error(f"Error al leer Excel con openpyxl: {e}")
        return {
            'success': False,
            'error': f"No se pudo leer el archivo Excel. Asegúrese de que sea un archivo .xlsx válido. ({str(e)})"
        }

    if not rows_data:
        return {'success': False, 'error': "El archivo está vacío."}

    # Normalizar encabezados (Fila 1)
    raw_headers = [str(h).strip() if h is not None else "" for h in rows_data[0]]
    
    # Mapeo flexible de encabezados
    header_map = {}
    for idx, h in enumerate(raw_headers):
        h_norm = h.lower().replace(" ", "_").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
        if "id" in h_norm and "actividad" in h_norm:
            header_map['id_actividad'] = idx
        elif "id" in h_norm and idx == 0:
            header_map['id_actividad'] = idx
        elif "titulo" in h_norm or "ponencia" in h_norm or "charla" in h_norm or "nombre" in h_norm:
            header_map['titulo'] = idx
        elif "inicio" in h_norm or "hora_ini" in h_norm:
            header_map['hora_inicio'] = idx
        elif "fin" in h_norm or "hora_fin" in h_norm:
            header_map['hora_fin'] = idx
        elif "aula" in h_norm or "sala" in h_norm or "escenario" in h_norm:
            header_map['aula'] = idx
        elif "categoria" in h_norm or "eje" in h_norm:
            header_map['categoria'] = idx
        elif "disertante" in h_norm or "ponente" in h_norm or "expositor" in h_norm:
            header_map['disertantes'] = idx
        elif "empresa" in h_norm or "institucion" in h_norm:
            header_map['empresas'] = idx
        elif "descripcion" in h_norm or "resumen" in h_norm:
            header_map['descripcion'] = idx
        elif "estado" in h_norm:
            header_map['estado'] = idx

    def get_raw_cell(row, key):
        if key in header_map and header_map[key] < len(row):
            return row[header_map[key]]
        return None

    def get_val(row, key, default=""):
        val = get_raw_cell(row, key)
        if val is None:
            return default
        val_str = str(val).strip()
        return "" if val_str in ('—', '-', 'None') else val_str

    def parse_time(val):
        if val is None or val == '' or str(val).strip() in ('—', '-', 'None'):
            return None
        if isinstance(val, time):
            return val
        if isinstance(val, datetime):
            return val.time()
        if isinstance(val, (int, float)):
            if 0.0 <= val <= 1.0:
                tot_sec = int(round(val * 86400))
                h = (tot_sec // 3600) % 24
                m = (tot_sec % 3600) // 60
                s = tot_sec % 60
                return time(h, m, s)
            elif val > 1.0:
                h = int(val)
                m = int(round((val - h) * 60))
                return time(h % 24, m % 60)

        val_str = str(val).strip()
        val_clean = val_str.replace('hs.', '').replace('hs', '').replace('hrs', '').replace('.', ':').strip()

        for fmt in ('%H:%M', '%H:%M:%S', '%I:%M %p', '%Y-%m-%d %H:%M:%S'):
            try:
                return datetime.strptime(val_clean, fmt).time()
            except ValueError:
                pass
        return None

    # Helper categorías y aulas
    categorias_validas = dict(Programa.CATEGORIA_CHOICES)
    aulas_validas = dict(Programa.AULA_CHOICES)

    filas_analizadas = []
    horarios_por_aula = {}  # Para detectar solapamientos dentro del mismo Excel

    # Parsear cada fila de datos (a partir de fila 2)
    for row_num, row_cells in enumerate(rows_data[1:], start=2):
        id_str = get_val(row_cells, 'id_actividad')
        id_actividad = int(id_str) if id_str and id_str.isdigit() else None

        titulo = get_val(row_cells, 'titulo')
        aula_raw = get_val(row_cells, 'aula', 'Aula Magna')
        aula = _mapear_aula(aula_raw)

        categoria_raw = get_val(row_cells, 'categoria', 'LOGISTICA')
        categoria = _mapear_categoria(categoria_raw)

        disertantes_str = get_val(row_cells, 'disertantes')
        empresas_str = get_val(row_cells, 'empresas')
        descripcion = get_val(row_cells, 'descripcion')
        estado_raw = get_val(row_cells, 'estado', 'PUBLICADO')
        estado = _mapear_estado(estado_raw)

        # Parsear tiempos desde la celda RAW
        raw_ini = get_raw_cell(row_cells, 'hora_inicio')
        raw_fin = get_raw_cell(row_cells, 'hora_fin')
        t_ini = parse_time(raw_ini)
        t_fin = parse_time(raw_fin)

        # Validación primaria
        errores_fila = []
        advertencias_fila = []

        if not titulo:
            errores_fila.append("El título de la actividad es obligatorio.")
        if not t_ini:
            errores_fila.append(f"La hora de inicio es inválida u obligatoria (valor recibido: '{raw_ini}').")
        if not t_fin:
            errores_fila.append(f"La hora de fin es inválida u obligatoria (valor recibido: '{raw_fin}').")
        if t_ini and t_fin and t_fin <= t_ini:
            errores_fila.append(f"La hora de fin ({t_fin.strftime('%H:%M')}) debe ser posterior a la de inicio ({t_ini.strftime('%H:%M')}).")

        if aula_raw and aula_raw != aula:
            advertencias_fila.append(f"Aula '{aula_raw}' normalizada como '{aula}'.")

        # Solapamiento interno en el Excel
        if t_ini and t_fin:
            key_aula = aula
            if key_aula not in horarios_por_aula:
                horarios_por_aula[key_aula] = []
            for prev_row, p_ini, p_fin in horarios_por_aula[key_aula]:
                if not (t_fin <= p_ini or t_ini >= p_fin):
                    errores_fila.append(f"Solapamiento horario en '{key_aula}' con la Fila {prev_row} ({p_ini.strftime('%H:%M')} - {p_fin.strftime('%H:%M')}).")
            horarios_por_aula[key_aula].append((row_num, t_ini, t_fin))

        # Disertantes advertencias
        if disertantes_str and disertantes_str not in ('—', '-', 'None'):
            # Separar disertantes por ';' o '/'
            lista_dis = [d.strip() for d in disertantes_str.replace('/', ';').split(";") if d.strip() and d.strip() not in ('—', '-')]
            for d_name in lista_dis:
                dis_db = Disertante.objects.filter(nombre__iexact=d_name, edicion=edicion).first()
                if not dis_db:
                    advertencias_fila.append(f"Se creará el perfil de nuevo disertante: '{d_name}'.")

        # Comparación contra Base de Datos (Delta Analysis)
        prog_existente = None
        if id_actividad:
            prog_existente = Programa.objects.filter(id=id_actividad, edicion=edicion).first()
        if not prog_existente and titulo:
            prog_existente = Programa.objects.filter(titulo__iexact=titulo, edicion=edicion).first()

        diff_map = {}
        estado_fila = 'NUEVO'

        if errores_fila:
            estado_fila = 'ERROR'
        elif prog_existente:
            # Comparar diferencias
            if prog_existente.titulo.strip() != titulo:
                diff_map['titulo'] = {'antes': prog_existente.titulo, 'ahora': titulo}
            if prog_existente.hora_inicio != t_ini:
                diff_map['hora_inicio'] = {
                    'antes': prog_existente.hora_inicio.strftime('%H:%M') if prog_existente.hora_inicio else 'Sin hora',
                    'ahora': t_ini.strftime('%H:%M') if t_ini else ''
                }
            if prog_existente.hora_fin != t_fin:
                diff_map['hora_fin'] = {
                    'antes': prog_existente.hora_fin.strftime('%H:%M') if prog_existente.hora_fin else 'Sin hora',
                    'ahora': t_fin.strftime('%H:%M') if t_fin else ''
                }
            if prog_existente.aula != aula:
                diff_map['aula'] = {'antes': prog_existente.aula, 'ahora': aula}
            if prog_existente.categoria != categoria:
                diff_map['categoria'] = {'antes': prog_existente.categoria, 'ahora': categoria}

            if diff_map:
                estado_fila = 'MODIFICADO'
            else:
                estado_fila = 'SIN_CAMBIOS'

        filas_analizadas.append({
            'num_fila': row_num,
            'id_actividad': prog_existente.id if prog_existente else id_actividad,
            'titulo': titulo,
            'hora_inicio': t_ini.strftime('%H:%M') if t_ini else str(raw_ini or ''),
            'hora_fin': t_fin.strftime('%H:%M') if t_fin else str(raw_fin or ''),
            'aula': aula,
            'categoria': categoria,
            'disertantes': disertantes_str,
            'empresas': empresas_str,
            'descripcion': descripcion,
            'estado': estado,
            'estado_fila': estado_fila,
            'diff': diff_map,
            'errores': errores_fila,
            'advertencias': advertencias_fila,
            '_t_ini': t_ini,
            '_t_fin': t_fin,
        })

    # Si commit = False, devolver análisis Pre-flight
    if not commit:
        resumen = {
            'total': len(filas_analizadas),
            'sin_cambios': sum(1 for f in filas_analizadas if f['estado_fila'] == 'SIN_CAMBIOS'),
            'nuevos': sum(1 for f in filas_analizadas if f['estado_fila'] == 'NUEVO'),
            'modificados': sum(1 for f in filas_analizadas if f['estado_fila'] == 'MODIFICADO'),
            'errores': sum(1 for f in filas_analizadas if f['estado_fila'] == 'ERROR'),
            'advertencias': sum(len(f['advertencias']) for f in filas_analizadas),
        }
        # Limpiar referencias internas no serializables en JSON
        for f in filas_analizadas:
            f.pop('_t_ini', None)
            f.pop('_t_fin', None)

        return {
            'success': True,
            'resumen': resumen,
            'filas': filas_analizadas
        }

    # Si commit = True: Ingesta atómica de filas aprobadas
    creados_count = 0
    actualizados_count = 0
    omitidos_count = 0

    aprobadas_set = set(filas_aprobadas) if filas_aprobadas is not None else None

    with transaction.atomic():
        for fila in filas_analizadas:
            row_num = fila['num_fila']
            # Si se especificaron filas aprobadas y esta fila no está aprobada, se omite
            if aprobadas_set is not None and row_num not in aprobadas_set:
                omitidos_count += 1
                continue
            if fila['estado_fila'] == 'ERROR':
                omitidos_count += 1
                continue
            if fila['estado_fila'] == 'SIN_CAMBIOS':
                omitidos_count += 1
                continue

            # Buscar o instanciar Programa
            prog_obj = None
            if fila['id_actividad']:
                prog_obj = Programa.objects.filter(id=fila['id_actividad'], edicion=edicion).first()
            if not prog_obj and fila['titulo']:
                prog_obj = Programa.objects.filter(titulo__iexact=fila['titulo'], edicion=edicion).first()

            if not prog_obj:
                prog_obj = Programa(edicion=edicion)
                created = True
            else:
                created = False

            prog_obj.titulo = fila['titulo']
            prog_obj.hora_inicio = fila['_t_ini']
            prog_obj.hora_fin = fila['_t_fin']
            prog_obj.dia = "2026-11-07"
            prog_obj.aula = fila['aula']
            prog_obj.categoria = fila['categoria']
            prog_obj.descripcion = fila['descripcion']
            prog_obj.estado = fila['estado']
            prog_obj.save()

            # Disertantes asociar/crear
            if fila['disertantes']:
                lista_nombres = [d.strip() for d in fila['disertantes'].split(";") if d.strip()]
                lista_empresas = [e.strip() for e in fila['empresas'].split(";") if e.strip()] if fila['empresas'] else []

                disertantes_objs = []
                for idx_d, d_name in enumerate(lista_nombres):
                    emp = lista_empresas[idx_d] if idx_d < len(lista_empresas) else ""
                    dis_obj = Disertante.objects.filter(nombre__iexact=d_name, edicion=edicion).first()
                    if not dis_obj:
                        dis_obj = Disertante.objects.create(
                            nombre=d_name,
                            empresa_institucion=emp,
                            edicion=edicion,
                            estado='APROBADO'
                        )
                    elif emp and not dis_obj.empresa_institucion:
                        dis_obj.empresa_institucion = emp
                        dis_obj.save()
                    disertantes_objs.append(dis_obj)

                prog_obj.disertantes.set(disertantes_objs)

            if created:
                creados_count += 1
            else:
                actualizados_count += 1

    return {
        'success': True,
        'creados': creados_count,
        'actualizados': actualizados_count,
        'omitidos': omitidos_count,
        'mensaje': f"Importación completada: {creados_count} creados, {actualizados_count} actualizados, {omitidos_count} omitidos."
    }


def exportar_excel_programa(edicion=None) -> HttpResponse:
    """Genera y descarga un archivo Excel (.xlsx) con el programa actual cargado en la base de datos."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if edicion is None:
        edicion = Edicion.objects.filter(activa=True).first()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Programa 2026"
    ws.views.sheetView[0].showGridLines = True

    headers = [
        "ID", "Título", "Hora Inicio", "Hora Fin", "Día", "Aula", "Categoría", "Estado", "Disertantes", "Empresas", "Descripción"
    ]
    ws.append(headers)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Consultar charlas
    qs = Programa.objects.all()
    if edicion:
        qs = qs.filter(edicion=edicion)
    qs = qs.order_by('aula', 'hora_inicio')

    for prog in qs:
        disertantes_str = "; ".join([d.nombre for d in prog.disertantes.all()])
        empresas_str = "; ".join([d.empresa_institucion for d in prog.disertantes.all() if d.empresa_institucion])
        
        row_data = [
            prog.id,
            prog.titulo,
            prog.hora_inicio.strftime("%H:%M") if prog.hora_inicio else "",
            prog.hora_fin.strftime("%H:%M") if prog.hora_fin else "",
            prog.dia.strftime("%Y-%m-%d") if prog.dia else "2026-11-07",
            prog.aula,
            prog.categoria,
            prog.estado,
            disertantes_str,
            empresas_str,
            prog.descripcion
        ]
        ws.append(row_data)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            if cell.column in (1, 3, 4, 5, 6, 7, 8):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_edicion = edicion.nombre.replace(" ", "_") if edicion else "2026"
    filename = f"Programa_Exportado_{nombre_edicion}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

